## 1. Import Required Libraries


```python
import os
import zipfile
import dask.dataframe as dd
import numpy as np
from pathlib import Path
import glob
import pandas as pd
from functools import reduce  
import shutil
```

## 2. Define Cell Mapping


```python
# Define your mappings
groups = {
    'Major Server': ['17625_Sunder Road Raiwind City Kasur (3G-CI-4984)',
                    '17181_Haveli Salamat Pura Kasur (3G-CI-1653)',
                    '17618_Railway Colony (Gold) Kasur (3G-CI-4981)',
                    '27180_Sunder Estate Kasur (3G-CI-1652)',
                    '33022_Pajian Village Lahore-Z3 (3G-CI-1560)',
                    'CI-1125-2_Block J Bahria Orchard Lahore (3G-CI-1166)',
                    'CI-1125-3_Block J Bahria Orchard Lahore (3G-CI-1166)',
                    'CI-1212-1_Near Masjid Khalid Bin Waleed (3G-CI-1234)',
                    'CI-1212-3_Near Masjid Khalid Bin Waleed (3G-CI-1234)',
                    '3G-CI-6560-3_Pajian Village Lahore-Z3 (3022)',
                    '3G-CI-9981-1_Railway Colony (Gold) Kasur (7618)',
                    '3G-CI-9984-1_Sunder Road Raiwind City Kasur (7625)',
                    '3G-CI-6652-2_Sunder Estate Kasur (7180)',
                    '3G-CI-6653-1_Haveli Salamat Pura Kasur (7181)',
                    '3G-CI-6166-2_Block J Bahria Orchard Lahore (CI-1125)',
                    '3G-CI-6166-3_Block J Bahria Orchard Lahore (CI-1125)',
                    '3G-CI-6234-3_Near Masjid Khalid Bin Waleed (CI-1212)',
                    '3G-CI-6234-1_Near Masjid Khalid Bin Waleed (CI-1212)',
                    '4G-CI-124984-1_Sunder Road Raiwind City Kasur (7625)',
                    '4G-CI-124984-81_Sunder Road Raiwind City Kasur (7625)',
                    '4G-CI-121653-1_Haveli Salamat Pura Kasur (7181)',
                    '4G-CI-121653-81_Haveli Salamat Pura Kasur (7181)',
                    '4G-CI-121560-3_Pajian Village Lahore-Z3 (3022)',
                    '4G-CI-121560-13_Pajian Village Lahore-Z3 (3022)',
                    '4G-CI-121652-2_Sunder Estate Kasur (7180)',
                    '4G-CI-124981-1_Railway Colony (Gold) Kasur (7618)',
                    '4G-CI-121652-12_Sunder Estate Kasur (7180)',
                    '4G-CI-124981-81_Railway Colony (Gold) Kasur (7618)',
                    '4G-CI-121166-3_Block J Bahria Orchard Lahore (CI-1125)',
                    '4G-CI-121166-2_Block J Bahria Orchard Lahore (CI-1125)',
                    '4G-CI-121234-3_Near Masjid Khalid Bin Waleed (CI-1212)',
                    '4G-CI-121234-1_Near Masjid Khalid Bin Waleed (CI-1212)',
                    '4G-CI-121234-83_Near Masjid Khalid Bin Waleed (CI-1212)',
                    '4G-CI-121234-81_Near Masjid Khalid Bin Waleed (CI-1212)'],
                
                'Other': ['37625_Sunder Road Raiwind City Kasur (3G-CI-4984)',
                '27625_Sunder Road Raiwind City Kasur (3G-CI-4984)',
                '27181_Haveli Salamat Pura Kasur (3G-CI-1653)',
                '37181_Haveli Salamat Pura Kasur (3G-CI-1653)',
                '27618_Railway Colony (Gold) Kasur (3G-CI-4981)',
                '37618_Railway Colony (Gold) Kasur (3G-CI-4981)',
                '17620_Main Chowk Raiwand Kasur (3G-CI-1746)',
                '27620_Main Chowk Raiwand Kasur (3G-CI-1746)',
                '37620_Main Chowk Raiwand Kasur (3G-CI-1746)',
                '37180_Sunder Estate Kasur (3G-CI-1652)',
                '17180_Sunder Estate Kasur (3G-CI-1652)',
                '13022_Pajian Village Lahore-Z3 (3G-CI-1560)',
                '23022_Pajian Village Lahore-Z3 (3G-CI-1560)',
                'CI-1125-1_Block J Bahria Orchard Lahore (3G-CI-1166)',
                'CI-1234-1_Jalalpura Sue-e-Asal Road (3G-CI-1253)',
                'CI-1234-2_Jalalpura Sue-e-Asal Road (3G-CI-1253)',
                'CI-1234-3_Jalalpura Sue-e-Asal Road (3G-CI-1253)',
                '3G-CI-6652-3_Sunder Estate Kasur (7180)',
                '3G-CI-6560-2_Pajian Village Lahore-Z3 (3022)',
                '3G-CI-6652-1_Sunder Estate Kasur (7180)',
                '3G-CI-6560-1_Pajian Village Lahore-Z3 (3022)',
                '3G-CI-6653-2_Haveli Salamat Pura Kasur (7181)',
                '3G-CI-6746-2_Main Chowk Raiwand Kasur (7620)',
                '3G-CI-9905-2_Raiwand City Kasur (3081)',
                '3G-CI-9981-2_Railway Colony (Gold) Kasur (7618)',
                '3G-CI-9983-2_Rehman Pura (Gold) Kasur (7621)',
                '3G-CI-9984-2_Sunder Road Raiwind City Kasur (7625)',
                '3G-CI-9983-1_Rehman Pura (Gold) Kasur (7621)',
                '3G-CI-6653-3_Haveli Salamat Pura Kasur (7181)',
                '3G-CI-6746-3_Main Chowk Raiwand Kasur (7620)',
                '3G-CI-9905-3_Raiwand City Kasur (3081)',
                '3G-CI-9981-3_Railway Colony (Gold) Kasur (7618)',
                '3G-CI-9983-3_Rehman Pura (Gold) Kasur (7621)',
                '3G-CI-9984-3_Sunder Road Raiwind City Kasur (7625)',
                '3G-CI-6746-1_Main Chowk Raiwand Kasur (7620)',
                '3G-CI-9905-1_Raiwand City Kasur (3081)',
                '3G-CI-6166-1_Block J Bahria Orchard Lahore (CI-1125)',
                '3G-CI-6253-3_Jalalpura Sue-e-Asal Road (CI-1234)',
                '3G-CI-6253-2_Jalalpura Sue-e-Asal Road (CI-1234)',
                '3G-CI-6253-1_Jalalpura Sue-e-Asal Road (CI-1234)',
                '3G-CI-6234-2_Near Masjid Khalid Bin Waleed (CI-1212)',
                '4G-CI-124984-3_Sunder Road Raiwind City Kasur (7625)',
                '4G-CI-124984-2_Sunder Road Raiwind City Kasur (7625)',
                '4G-CI-124984-82_Sunder Road Raiwind City Kasur (7625)',
                '4G-CI-124984-83_Sunder Road Raiwind City Kasur (7625)',
                '4G-CI-124983-3_Rehman Pura (Gold) Kasur (7621)',
                '4G-CI-124983-2_Rehman Pura (Gold) Kasur (7621)',
                '4G-CI-124983-1_Rehman Pura (Gold) Kasur (7621)',
                '4G-CI-121746-1_Main Chowk Raiwand Kasur (7620)',
                '4G-CI-121746-2_Main Chowk Raiwand Kasur (7620)',
                '4G-CI-121746-3_Main Chowk Raiwand Kasur (7620)',
                '4G-CI-121653-3_Haveli Salamat Pura Kasur (7181)',
                '4G-CI-121653-2_Haveli Salamat Pura Kasur (7181)',
                '4G-CI-121746-81_Main Chowk Raiwand Kasur (7620)',
                '4G-CI-121746-82_Main Chowk Raiwand Kasur (7620)',
                '4G-CI-121746-83_Main Chowk Raiwand Kasur (7620)',
                '4G-CI-124983-81_Rehman Pura (Gold) Kasur (7621)',
                '4G-CI-124983-82_Rehman Pura (Gold) Kasur (7621)',
                '4G-CI-124983-83_Rehman Pura (Gold) Kasur (7621)',
                '4G-CI-121653-82_Haveli Salamat Pura Kasur (7181)',
                '4G-CI-121653-83_Haveli Salamat Pura Kasur (7181)',
                '4G-CI-124905-3_Raiwand City Kasur (3081)',
                '4G-CI-124905-2_Raiwand City Kasur (3081)',
                '4G-CI-124905-1_Raiwand City Kasur (3081)',
                '4G-CI-124905-83_Raiwand City Kasur (3081)',
                '4G-CI-124905-81_Raiwand City Kasur (3081)',
                 '4G-CI-124905-82_Raiwand City Kasur (3081)',
                '4G-CI-121560-2_Pajian Village Lahore-Z3 (3022)',
                '4G-CI-121560-1_Pajian Village Lahore-Z3 (3022)',
                '4G-CI-121652-3_Sunder Estate Kasur (7180)',
                '4G-CI-121652-1_Sunder Estate Kasur (7180)',
                '4G-CI-124981-3_Railway Colony (Gold) Kasur (7618)',
                '4G-CI-124981-2_Railway Colony (Gold) Kasur (7618)',
                '4G-CI-124981-82_Railway Colony (Gold) Kasur (7618)',
                '4G-CI-124981-83_Railway Colony (Gold) Kasur (7618)',
                '4G-CI-121253-3_Jalalpura Sue-e-Asal Road (CI-1234)',
                '4G-CI-121253-2_Jalalpura Sue-e-Asal Road (CI-1234)',
                '4G-CI-121253-1_Jalalpura Sue-e-Asal Road (CI-1234)',
                '4G-CI-121166-1_Block J Bahria Orchard Lahore (CI-1125)',
                '4G-CI-121234-2_Near Masjid Khalid Bin Waleed (CI-1212)',
                '4G-CI-121234-82_Near Masjid Khalid Bin Waleed (CI-1212)',
                ],
    
                'Raiwind': ['13081_Raiwand City Kasur (3G-CI-4905)',
                '23081_Raiwand City Kasur (3G-CI-4905)',
                '33081_Raiwand City Kasur (3G-CI-4905)',
                'CI-1212-2_Near Masjid Khalid Bin Waleed (3G-CI-1234)',
                '17621_Rehman Pura (Gold) Kasur (3G-CI-4983)',
                '27621_Rehman Pura (Gold) Kasur (3G-CI-4983)',
                '37621_Rehman Pura (Gold) Kasur (3G-CI-4983)']
                                }

# Convert to a flat DataFrame
groups_df = pd.DataFrame([(name, degree) for degree, names in groups.items() for name in names],
                  columns=['Cell Name', 'Zone'])

groups_df
```

## 3. Load & Extract Data from ZIP Files


```python
def load_from_zip_folder(zip_folder, keyword, required_columns, dtype=None):
    """
    🗜️ Extracts and combines CSV files from ZIPs in a folder, filtered by keyword and required columns.

    Parameters:
        zip_folder (str or Path): 📁 Folder containing ZIP files.
        keyword (str): 🔍 Keyword to filter relevant CSV filenames (e.g., "(GSM_Cell_Hourly)").
        required_columns (list): 🧱 List of required columns to read from CSVs.
        dtype (dict): Optional 🧬 dictionary specifying data types for specific columns.

    Returns:
        pd.DataFrame: 🐼 Combined DataFrame with valid rows, or None if no match found.
    """
    zip_files = list(Path(zip_folder).glob("*.zip"))  # 📦 List all ZIP files in the folder
    dfs = []  # 📋 Initialize list to collect DataFrames

    for zip_file in zip_files:
        with zipfile.ZipFile(zip_file, 'r') as z:
            # 🎯 Filter CSV files inside the ZIP that match the keyword
            csv_files = [f for f in z.namelist() if keyword in f and f.endswith(".csv")]

            for csv_file in csv_files:
                with z.open(csv_file) as f:
                    try:
                        df = pd.read_csv(
                            f,
                            usecols=required_columns,      # 🧩 Only load required columns
                            dtype=dtype,                   # 🧬 Apply custom dtypes if given
                            parse_dates=["Date"],          # 📆 Ensure 'Date' is parsed as datetime
                            skiprows=range(6),             # 🪄 Skip first 6 header rows (formatting)
                            skipfooter=1,                  # 🚫 Skip footer line if present
                            engine='python',               # 🐍 Use Python engine for flexibility
                            na_values=['NIL', '/0']        # ❓ Replace invalid entries with NaN
                        )
                        dfs.append(df)  # ✅ Add valid DataFrame to the list
                    except ValueError:
                        # ⚠️ Handle files missing required columns gracefully
                        print(f"Skipping {csv_file} in {zip_file.name} - Missing required columns.")

    if dfs:
        return pd.concat(dfs, ignore_index=True)  # 🧩 Combine all valid DataFrames
    else:
        print("No matching CSV files found or required columns missing. ❌")
        return None
```

## 4. Site/Cell ID Generater


```python
def extract_lte_ids(ddf, column_map):
    for source_col, new_col in column_map.items():
        ddf[new_col] = ddf[source_col].str.extract(r'^([^_]+)', expand=False)
    return ddf
```

## 5. Filter Max Data and Define Max Hours


```python
def filter_max_date_last_n_hours(df, date_col='Date', time_col='Time', n=8):
    """
    Filter the DataFrame for the maximum date, 
    then keep only rows from the last n hours of that date.
    """
    df = df.copy()

    # Convert date and time columns
    df[date_col] = pd.to_datetime(df[date_col])
    df[time_col] = pd.to_datetime(df[time_col], format='%H:%M', errors='coerce')

    # 1️⃣ Find maximum date
    max_date = df[date_col].max()

    # 2️⃣ Filter for that date
    df_max = df[df[date_col] == max_date].copy()

    # 3️⃣ Sort by Time and find the cutoff for last n hours
    df_max = df_max.sort_values(by=time_col)
    unique_times = df_max[time_col].drop_duplicates().sort_values()

    if len(unique_times) > n:
        cutoff_time = unique_times.iloc[-n]
        df_last_nh = df_max[df_max[time_col] >= cutoff_time].copy()
    else:
        df_last_nh = df_max.copy()  # if fewer than n hours, return all

    # 4️⃣ Safely format Time column (avoids SettingWithCopyWarning)
    df_last_nh.loc[:, time_col] = df_last_nh[time_col].dt.strftime('%H:%M')

    return df_last_nh.reset_index(drop=True)
```

## 6. Load GSM Data & Process required data


```python
#  Define path and parameters for GSM Cell Hourly data
zip_folder = "D:/Advance_Data_Sets/Raiwind"
keyword_gsm = "(2G)"
required_columns_gsm = [
    'Date','Time','Cell Name',
    'GOS-SDCCH(%)','CallSetup TCH GOS(%)'
]

#  Load GSM data from ZIP files 
gsm_df = load_from_zip_folder(zip_folder, keyword_gsm, required_columns_gsm)

# Generate GSM Cell ID
gsm_df = extract_lte_ids(gsm_df, column_map={'Cell Name': 'Cell ID'})

# merge the GSM Cell Hourly data with groups to get the zone
gsm_df_map = pd.merge(gsm_df,groups_df,how='left',on=['Cell Name'])

# filter max date and define max hrs
gsm_df_max_date_hr = filter_max_date_last_n_hours(gsm_df_map, date_col='Date', time_col='Time')
```


```python
gsm_df_max_date_hr
```

## 7. Load UMTS Data & Process required data


```python
#  Define path and parameters for UMTS Cell Hourly data
keyword_umts = "(3G)"
required_columns_umts = [
    'Date','Time','RNC','Cell Name','CS RAB Congestion Ratio (%)'
]

#  Load UMTS data from ZIP files 
umts_df = load_from_zip_folder(zip_folder, keyword_umts, required_columns_umts)

# Generate UMTS Cell ID
umts_df = extract_lte_ids(umts_df, column_map={'Cell Name': 'Cell ID'})

# merge the UMTS Cell Hourly data with groups to get the zone
umts_df_map = pd.merge(umts_df,groups_df,how='left',on=['Cell Name'])

# filter max date and define max hrs
umts_df_max_date_hr = filter_max_date_last_n_hours(umts_df_map, date_col='Date', time_col='Time')
umts_df_max_date_hr
```

## 8. Load LTE Data & Process required data


```python
#  Define path and parameters for LTE Cell Hourly data
keyword_lte = "(4G)"

required_columns_lte = [
    'Date','Time','Cell Name','Data Volume (GB)','DL User Thrp (Mbps)','Data Volume (GB)',
]

#  Load LTE data from ZIP files 
lte_df = load_from_zip_folder(zip_folder, keyword_lte, required_columns_lte)

# Generate LTE Cell ID
lte_df = extract_lte_ids(lte_df, column_map={'Cell Name': 'Cell ID'})

# merge the LTE Cell Hourly data with groups to get the zone
lte_df_map = pd.merge(lte_df,groups_df,how='left',on=['Cell Name'])

# filter max date and define max hrs
lte_df_max_date_hr = filter_max_date_last_n_hours(lte_df_map, date_col='Date', time_col='Time')
lte_df_max_date_hr
```

## 9. Required pivot tables


```python
pivot_sources = {
    'gsm': gsm_df_max_date_hr,
    'umts': umts_df_max_date_hr,
    'lte': lte_df_max_date_hr
}

pivot_metrics = {
    'CS_TCH_GoS': ('gsm', 'CallSetup TCH GOS(%)'),
    'GoS_SDCCH': ('gsm', 'GOS-SDCCH(%)'),
    'CS_RAB_Cong_Ratio': ('umts', 'CS RAB Congestion Ratio (%)'),
    'DL_Thrp': ('lte', 'DL User Thrp (Mbps)'),
    'Data_Volume' : ('lte','Data Volume (GB)')
}


pivots = {
    key: pd.pivot_table(
        pivot_sources[src],
        index=['Date', 'Cell ID', 'Zone'],
        columns='Time',
        values=metric
    )
    .reset_index()
    .assign(Comment="")  # 👈 Adds an empty Comment column
    for key, (src, metric) in pivot_metrics.items()
}

# Access them like this:
CS_TCH_GoS, GoS_SDCCH, CS_RAB_Cong_Ratio, DL_Thrp,Data_Volume = [
    pivots[k] for k in ['CS_TCH_GoS', 'GoS_SDCCH', 'CS_RAB_Cong_Ratio', 'DL_Thrp','Data_Volume']]



```

## 10. Export Funcation 


```python
def export_zone_dataframes(zone_name, output_file):
    """
    Export all pivot DataFrames filtered by a specific Zone
    into a single Excel file with separate sheets.
    """
    pivots_dict = {
        'CS_TCH_GoS': CS_TCH_GoS,
        'GoS_SDCCH': GoS_SDCCH,
        'CS_RAB_Cong_Ratio': CS_RAB_Cong_Ratio,
        'DL_Thrp': DL_Thrp,
        'Data_Volume': Data_Volume
        
    }

    with pd.ExcelWriter(output_file) as writer:
        for name, df in pivots_dict.items():
            # Filter and make a copy
            if 'Zone' in df.columns:
                df_filtered = df[df['Zone'] == zone_name].copy()
            else:
                df_filtered = df.copy()

            # ✅ Convert Date column to string (prevents 00:00:00 issue)
            if 'Date' in df_filtered.columns:
                df_filtered['Date'] = pd.to_datetime(df_filtered['Date']).dt.strftime('%Y-%m-%d').astype(str)

            # Format time-based column headers (if any)
            df_filtered.columns = df_filtered.columns.map(
                lambda x: x.strftime('%H:%M') if isinstance(x, pd.Timestamp) else x
            )

            # Export to Excel
            df_filtered.to_excel(writer, index=False, sheet_name=name, startrow=2,startcol=0)
```

## 11. File Path Setting


```python
# Set Working Folder Path
path = 'D:/Advance_Data_Sets/Raiwind/Output'

# Check if folder exists
if not os.path.exists(path):
    os.makedirs(path)  # Create the folder if it doesn't exist
    print(f"📁 Folder created: {path}")
else:
    # Folder exists → remove all files inside it
    for filename in os.listdir(path):
        file_path = os.path.join(path, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.remove(file_path)  # Delete file
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)  # Delete subfolder
        except Exception as e:
            print(f"⚠️ Failed to delete {file_path}. Reason: {e}")
    print(f"🧹 All files deleted in: {path}")

# Set as working directory
os.chdir(path)
print(f"✅ Working directory set to: {os.getcwd()}")
```

## 12. Export


```python
export_zone_dataframes('Major Server', 'Major.xlsx')
export_zone_dataframes('Other', 'Other.xlsx')
export_zone_dataframes('Raiwind', 'Raiwind.xlsx')
```


```python
#re-set all the variable from the RAM
%reset -f
```

## 13. Format Table


```python
import os
import openpyxl
from openpyxl import load_workbook
```


```python
# set the input file path
os.chdir('D:/Advance_Data_Sets/Raiwind/Output')
```


```python
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import openpyxl


# ======================================================
# STEP 1️⃣: Define and Register Styles
# ======================================================
def define_styles(wb):
    """Define and register custom styles for workbook."""
    font_style = Font(name='Calibri Light', size=11)
    alignment_style = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # General style
    style_name = "styled_cell"
    if style_name not in wb.named_styles:
        style = NamedStyle(name=style_name)
        style.font = font_style
        style.alignment = alignment_style
        style.border = thin_border
        wb.add_named_style(style)

    # Date style
    date_style_name = "custom_date_style"
    if date_style_name not in wb.named_styles:
        date_style = NamedStyle(name=date_style_name, number_format='DD/MM/YYYY')
        date_style.font = font_style
        date_style.alignment = alignment_style
        date_style.border = thin_border
        wb.add_named_style(date_style)

    return style_name, date_style_name


# ======================================================
# STEP 2️⃣: Format Header Row
# ======================================================
def format_headers(ws):
    """Format header row (Row 3) with color, bold text, and border."""
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12, name='Calibri Light')
    header_alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border


# ======================================================
# STEP 3️⃣: Apply Styles to Data Cells
# ======================================================
def apply_data_styles(ws, style_name, date_style_name):
    """Apply border, alignment, and number formatting to all data cells."""
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.style = date_style_name
            else:
                cell.style = style_name
            if isinstance(cell.value, float):
                cell.number_format = '0.00'


# ======================================================
# STEP 4️⃣: Auto Adjust Column Widths
# ======================================================
def auto_adjust_column_widths(ws):
    """Automatically adjust column widths based on content length."""
    for column in ws.columns:
        col_letter = column[0].column_letter
        max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column)
        ws.column_dimensions[col_letter].width = max_length + 8


# ======================================================
# STEP 5️⃣: Convert Specific Columns to Float
# ======================================================
def convert_columns_to_float(ws, columns):
    """Convert specific columns from string to float, if possible."""
    for column_letter in columns:
        column_range = ws[column_letter]
        for cell in column_range:
            if isinstance(cell.value, str):
                try:
                    cell.value = float(cell.value)
                except ValueError:
                    pass


# ======================================================
# STEP 6️⃣-A: Apply Sheet Properties (Zoom, Color, Filter)
# ======================================================
def apply_sheet_properties(ws, color, zoom):
    """Set zoom, tab color, and auto-filter."""
    ws.sheet_properties.tabColor = color
    ws.sheet_view.zoomScale = zoom
    if ws.max_row >= 3 and ws.max_column >= 1:
        last_col_letter = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A3:{last_col_letter}3"


# ======================================================
# STEP 6️⃣-B: Format a Single Sheet
# ======================================================
def format_single_sheet(ws, color, zoom, style_name, date_style_name):
    """Apply all formatting steps to a single sheet."""
    apply_sheet_properties(ws, color, zoom)
    format_headers(ws)
    apply_data_styles(ws, style_name, date_style_name)
    convert_columns_to_float(ws, ['B'])  # Example column
    auto_adjust_column_widths(ws)


# ======================================================
# STEP 6️⃣-C: Format Entire Workbook
# ======================================================
def format_workbook(file_path, colors=None, zoom=80):
    """Apply complete formatting to the entire workbook."""
    if colors is None:
        colors = ["00B0F0", "0000FF", "ADD8E6", "87CEFA"]

    wb = load_workbook(file_path)
    style_name, date_style_name = define_styles(wb)

    wb.calculation.calcMode = 'manual'
    for i, ws in enumerate(wb.worksheets):
        format_single_sheet(ws, colors[i % len(colors)], zoom, style_name, date_style_name)
    wb.calculation.calcMode = 'auto'

    wb.save(file_path)
    print(f"✅ Formatting applied to workbook: {file_path}")


# ======================================================
# STEP 7️⃣: Create Title Page
# ======================================================
from openpyxl.drawing.image import Image
import openpyxl

def add_title_page(file_path):
    """Add a custom title page with logos and formatted title."""
    wb = openpyxl.load_workbook(file_path)
    ws511 = wb.create_sheet("Home", 0)

    # ---- Insert Huawei logo ----
    try:
        img_huawei = Image('D:/Advance_Data_Sets/KPIs_Analysis/Huawei.jpg')
        img_huawei.width = 7 * 15
        img_huawei.height = 7 * 15
        ws511.add_image(img_huawei, 'E3')
    except FileNotFoundError:
        print("⚠️ Huawei logo not found at the specified path.")

    # ---- Insert PTCL logo ----
    # try:
    #     img_ptcl = Image('D:/Advance_Data_Sets/KPIs_Analysis/PTCL.png')
    #     img_ptcl.width = 7 * 15
    #     img_ptcl.height = 7 * 15
    #     ws511.add_image(img_ptcl, 'M3')
    # except FileNotFoundError:
    #     print("⚠️ PTCL logo not found at the specified path.")

    # ---- Create merged title area ----
    ws511.merge_cells(start_row=12, start_column=5, end_row=18, end_column=17)
    ws511.cell(row=12, column=5).value = 'Raiwind IJtema KPIs'

    # ---- Apply styling to the title ----
    first_row = list(ws511.rows)[11]
    for cell in first_row[4:]:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        cell.fill = openpyxl.styles.PatternFill(start_color="CF0A2C", end_color="CF0A2C", fill_type="solid")
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True, size=70, name='Calibri Light')

   

    wb.save(file_path)
    print("✅ Title Page added successfully.")

## Step 9️⃣ — Hide Gridlines and Headers (All Sheets)

def hide_gridlines_and_headers(file_path):
    """Hide gridlines and row/column headers for all sheets in the workbook."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.showRowColHeaders = False
    wb.save(file_path)
    print("✅ Gridlines and row/column headers hidden for all sheets.")

## Step 10- hyperlink
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def add_hyperlinks(file_path):
    """Add hyperlinks to all sheets and create navigation links in each sheet."""
    wb = load_workbook(file_path)
    
    # Define border style
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Reference the title (Home) sheet
    ws_home = wb["Home"]
    
    # === (1) Add TOC links in Home sheet ===
    row = 22
    for sheet in wb.worksheets:
        if sheet.title != "Home":
            hyperlink_cell = ws_home.cell(row=row, column=5)
            hyperlink_cell.value = sheet.title
            hyperlink_cell.hyperlink = f"#'{sheet.title}'!A1"
            hyperlink_cell.font = Font(color="0000FF", underline="single")
            hyperlink_cell.alignment = Alignment(horizontal='center')
            
            # Set row height and column width
            ws_home.row_dimensions[row].height = 15
            ws_home.column_dimensions[get_column_letter(5)].width = 30
            row += 1

    # === (2) Add “Back to TOC” and navigation links in each sheet ===
    for i, sheet in enumerate(wb.worksheets):
        if sheet.title != "Home":
            # Back to TOC link
            backlink_cell = sheet.cell(row=3, column=sheet.max_column + 2)
            backlink_cell.value = "Home"
            backlink_cell.hyperlink = "#'Home'!E22"
            backlink_cell.font = Font(color="0000FF", underline="single")
            backlink_cell.alignment = Alignment(horizontal='center')

            # Border and column width
            # for row in sheet.iter_rows(min_row=4, max_row=4,
            #                            min_col=backlink_cell.column,
            #                            max_col=backlink_cell.column):
            #     for cell in row:
            #         cell.border = border

            col_letter = get_column_letter(backlink_cell.column)
            sheet.column_dimensions[col_letter].width = 25

            # Next Sheet link
            if i < len(wb.worksheets) - 1:
                next_cell = sheet.cell(row=5, column=sheet.max_column)
                next_cell.value = "Next Sheet"
                next_cell.hyperlink = f"#'{wb.worksheets[i+1].title}'!A1"
                next_cell.font = Font(color="0000FF", underline="single")
                next_cell.alignment = Alignment(horizontal='center')

            # Previous Sheet link
            if i > 0:
                prev_cell = sheet.cell(row=4, column=sheet.max_column)
                prev_cell.value = "Previous Sheet"
                prev_cell.hyperlink = f"#'{wb.worksheets[i-1].title}'!A1"
                prev_cell.font = Font(color="0000FF", underline="single")
                prev_cell.alignment = Alignment(horizontal='center')

    wb.save(file_path)


##--------------------------------
from openpyxl import load_workbook

def set_home_as_active(file_path):
    """📘 Set 'Home' (Title Page) as the active sheet and deselect others."""
    wb = load_workbook(file_path)

    # Deselect all sheets
    for sheet in wb.worksheets:
        if hasattr(sheet, "views") and sheet.views and sheet.views.sheetView:
            sheet.views.sheetView[0].tabSelected = False

    # Set 'Home' as active
    if 'Home' in wb.sheetnames:
        wb.active = wb.sheetnames.index('Home')
        ws_home = wb['Home']
        if hasattr(ws_home, "views") and ws_home.views and ws_home.views.sheetView:
            ws_home.views.sheetView[0].tabSelected = True

    wb.save(file_path)
    print("🏠 'Home' sheet set as active and all others deselected.")

# ======================================================
# STEP 8️⃣: Master Function — Apply All
# ======================================================
def apply_excel_formatting(file_path):
    """Master function that adds title page and applies all formatting."""
    
    format_workbook(file_path)
    add_title_page(file_path)
    hide_gridlines_and_headers(file_path)
    add_hyperlinks(file_path)
    set_home_as_active(file_path)
    print("🎯 Complete Excel formatting applied (Title Page + Workbook Styling).")
```


```python
apply_excel_formatting(r"D:\Advance_Data_Sets\Raiwind\Output\Major.xlsx")
apply_excel_formatting(r"D:\Advance_Data_Sets\Raiwind\Output\Other.xlsx")
apply_excel_formatting(r"D:\Advance_Data_Sets\Raiwind\Output\Raiwind.xlsx")
```


```python
#re-set all the variable from the RAM
%reset -f
```

## 14. Home Sheet


```python
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

def add_sheet_name_title(file_path):
    """🧾 Add sheet name in A2:J2 (merged, centered, black background, white bold text),
    skipping the sheet named 'Home'."""
    
    wb = load_workbook(file_path)

    for ws in wb.worksheets:
        # ✅ Skip only the sheet named exactly "Home"
        if ws.title == "Home":
            continue

        merge_range = "A2:L2"

        # Merge cells A2 to J2
        ws.merge_cells(merge_range)

        # Write sheet name in A2
        cell = ws["A2"]
        cell.value = ws.title

        # Center align text horizontally and vertically
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply black fill
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        # Apply white bold font
        cell.font = Font(color="FFFFFF", bold=True)

    # Save the workbook
    wb.save(file_path)
    print("✅ Sheet names added (A2:J2) with black background and white bold text — except 'Home' sheet.")

# ✅ Example usage
add_sheet_name_title(r"D:\Advance_Data_Sets\Raiwind\Output\Major.xlsx")
add_sheet_name_title(r"D:\Advance_Data_Sets\Raiwind\Output\Other.xlsx")
add_sheet_name_title(r"D:\Advance_Data_Sets\Raiwind\Output\Raiwind.xlsx")
```


```python
#re-set all the variable from the RAM
%reset -f
```

## 15. Conditional Formatting


```python
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def apply_conditional_formatting_file(file_path):
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    wb = load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Determine threshold and comparison type
        if sheet_name.lower() == 'cs_tch_gos':
            threshold = 2
            compare_greater = True
        elif sheet_name.lower() == 'gos_sdcch':
            threshold = 0.1
            compare_greater = True
        elif sheet_name.lower() == 'cs_rab_cong_ratio':
            threshold = 2
            compare_greater = True
        elif sheet_name.lower() == 'dl_thrp':
            threshold = 1
            compare_greater = False  # Highlight if less than 1
        else:
            continue
        
        # Iterate rows and columns from column 4 (D) onward
        for row in ws.iter_rows(min_col=4, max_col=ws.max_column):
            for cell in row:
                try:
                    if cell.value is not None:
                        value = float(cell.value)
                        if compare_greater and value > threshold:
                            cell.fill = red_fill
                        elif not compare_greater and value < threshold:
                            cell.fill = red_fill
                except (ValueError, TypeError):
                    continue
    
    wb.save(file_path)
    print(f"Formatted: {file_path}")

# Example usage for single files
apply_conditional_formatting_file(r"D:\Advance_Data_Sets\Raiwind\Output\Major.xlsx")
apply_conditional_formatting_file(r"D:\Advance_Data_Sets\Raiwind\Output\Other.xlsx")
apply_conditional_formatting_file(r"D:\Advance_Data_Sets\Raiwind\Output\Raiwind.xlsx")
```


```python
#re-set all the variable from the RAM
%reset -f
```

